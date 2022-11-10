# -*- coding:utf-8 -*-
# @Time: 2022/10/17 0:04
# @Author: jooh
# @File: handle_excel.py
#    封装Excel的类


import openpyxl


class HandleExcel:
    """
    封装Excel的类
    """

    def __init__(self, filename, sheetname):
        """
        初始化方法，读写都需要包括文件名和表单名，因此提取出来
        :param filename:
        :param sheetname:
        """
        self.filename = filename
        self.sheetname = sheetname

    def read_data(self):
        """
        读取excel数据
        :return:
        """
        wb = openpyxl.load_workbook(self.filename)
        sh = wb[self.sheetname]
        res = list(sh.rows)
        # 获取第一行表头
        title = [i.value for i in res[0]]
        cases = []
        # 遍历第一行之外的其他行
        for item in res[1:]:
            data = [i.value for i in item]
            dic = dict(zip(title, data))
            cases.append(dic)
        return cases

    def write_excel(self, row, column, value):
        """
        写入excel数据
        :param row:
        :param column:
        :param value:
        :return:
        """
        wb = openpyxl.load_workbook(self.filename)
        sh = wb[self.sheetname]
        # 数据写入，这个步骤只是将内容写到工作簿中
        sh.cell(row=row, column=column, value=value)
        # 要保存才能写到文件中
        wb.save(self.filename)

